"""Base report class with export helpers (PDF, Excel)."""

from __future__ import annotations

import io
from typing import Any

from pydantic import BaseModel


class BaseReport(BaseModel):
    """Base class สำหรับ report ทุกประเภท — มี export methods."""

    model_config = {"arbitrary_types_allowed": True}

    def to_excel(self, sheet_name: str = "Report") -> bytes:
        """Export เป็น Excel (.xlsx) ด้วย openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
        header_font_white = Font(bold=True, color="FFFFFF")

        data = self.model_dump()
        self._write_dict_to_sheet(ws, data, header_font=header_font_white, header_fill=header_fill)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _write_dict_to_sheet(self, ws, data: dict, row: int = 1, indent: int = 0,
                              header_font=None, header_fill=None) -> int:
        """เขียน dict ลง worksheet แบบ recursive."""
        try:
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return row

        for key, value in data.items():
            if isinstance(value, list):
                ws.cell(row=row, column=indent + 1, value=str(key).upper())
                if header_font:
                    ws.cell(row=row, column=indent + 1).font = header_font
                if header_fill:
                    ws.cell(row=row, column=indent + 1).fill = header_fill
                row += 1
                for item in value:
                    if isinstance(item, dict):
                        for col_idx, (k, v) in enumerate(item.items(), start=indent + 1):
                            ws.cell(row=row, column=col_idx, value=str(v) if v is not None else "")
                        row += 1
                    else:
                        ws.cell(row=row, column=indent + 1, value=str(item))
                        row += 1
            elif isinstance(value, dict):
                ws.cell(row=row, column=indent + 1, value=str(key).upper())
                row += 1
                row = self._write_dict_to_sheet(ws, value, row=row, indent=indent + 1)
            else:
                ws.cell(row=row, column=indent + 1, value=str(key))
                ws.cell(row=row, column=indent + 2, value=value)
                row += 1
        return row

    def to_pdf(self, title: str = "Report", css: str | None = None) -> bytes:
        """Export เป็น PDF ด้วย WeasyPrint."""
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            raise ImportError("weasyprint is required: pip install weasyprint")

        html = self._to_html(title)
        default_css = """
        @page { size: A4; margin: 2cm; }
        body { font-family: 'Sarabun', sans-serif; font-size: 10pt; }
        h1 { color: #1F3864; font-size: 14pt; }
        h2 { color: #2E75B6; font-size: 11pt; border-bottom: 1px solid #ccc; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 1em; }
        th { background: #4472C4; color: white; padding: 4px 8px; text-align: left; }
        td { padding: 3px 8px; border-bottom: 1px solid #eee; }
        .number { text-align: right; }
        .total { font-weight: bold; background: #f0f0f0; }
        .negative { color: red; }
        """
        return HTML(string=html).write_pdf(stylesheets=[CSS(string=css or default_css)])

    def _to_html(self, title: str) -> str:
        """สร้าง HTML สำหรับ PDF export — override ใน subclass เพื่อ custom layout."""
        data = self.model_dump()
        rows = self._dict_to_html_rows(data)
        return f"""<html><head><meta charset='utf-8'></head>
<body><h1>{title}</h1>{rows}</body></html>"""

    def _dict_to_html_rows(self, data: dict, level: int = 0) -> str:
        html = ""
        for key, value in data.items():
            if isinstance(value, list):
                html += f"<h{min(level+2,4)}>{key}</h{min(level+2,4)}>"
                if value and isinstance(value[0], dict):
                    cols = list(value[0].keys())
                    html += "<table><tr>" + "".join(f"<th>{c}</th>" for c in cols) + "</tr>"
                    for item in value:
                        html += "<tr>" + "".join(
                            f"<td class='number'>{item.get(c,'')}</td>" for c in cols
                        ) + "</tr>"
                    html += "</table>"
            elif isinstance(value, dict):
                html += f"<h{min(level+2,4)}>{key}</h{min(level+2,4)}>"
                html += self._dict_to_html_rows(value, level + 1)
            else:
                html += f"<p><strong>{key}:</strong> {value}</p>"
        return html
